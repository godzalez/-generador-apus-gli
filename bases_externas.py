"""
bases_externas.py — Motor de búsqueda en bases de precios externas.
GLI Colombia S.A.S. · v12.0 · 2026

Fuentes soportadas (en orden de cascada):
  1. APU de la entidad contratante (Camino A — generator.py)
  2. Catálogo Policía Nacional 1LF-FR-0206 (zonas A1-A6)
  3. APU INVIAS regionalizados (Excel descargado por provincia)
  4. Cualquier base externa genérica (gobernación, IDU, etc.)
"""
import unicodedata, re
import pandas as pd
from openpyxl import load_workbook


# ══════════════════════════════════════════════════════════════════
# NORMALIZACIÓN
# ══════════════════════════════════════════════════════════════════

def _norm(texto):
    s = str(texto).upper().strip()
    s = ''.join(c for c in unicodedata.normalize('NFD', s)
                if unicodedata.category(c) != 'Mn')
    return re.sub(r'\s+', ' ', s)

_STOPWORDS = {
    'DE','LA','EL','EN','CON','Y','A','E','O','UN','UNA','LOS','LAS',
    'DEL','AL','POR','SU','ES','SE','QUE','PARA','HASTA','DESDE','SIN',
    'NO','SI','NI','MAS','TIPO','SEGUN','INCLUYE',
}


# ══════════════════════════════════════════════════════════════════
# DETECTORES DE FORMATO
# ══════════════════════════════════════════════════════════════════

def _detectar_formato(ws):
    primera = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    cols = [str(v).upper().strip() if v else '' for v in primera]
    cols_str = ' '.join(cols)

    if ('CODIGO ITEM' in cols_str and 'UNIDAD' in cols_str
            and 'TOTAL' in cols_str and 'CAPITULO' in cols_str):
        return 'boyaca'

    # INVIAS: columnas ITEM, DESCRIPCION_ACTIVIDAD, UNIDAD, COSTO_DIRECTO
    if ('DESCRIPCION_ACTIVIDAD' in cols_str or
            ('ITEM' in cols_str and 'COSTO_DIRECTO' in cols_str)):
        return 'invias'

    if any(x in cols_str for x in ('CODIGO', 'ITEM', 'DESCRIPCION')) and \
       any(x in cols_str for x in ('PRECIO', 'VALOR', 'TOTAL', 'UNITARIO')):
        return 'generico'

    return None


# ══════════════════════════════════════════════════════════════════
# CARGADORES POR FORMATO
# ══════════════════════════════════════════════════════════════════

def _cargar_boyaca(ws):
    bd = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[4] or not row[8]: continue
        total = float(row[8]) if isinstance(row[8], (int, float)) else 0
        if total <= 0: continue
        mo = float(row[7]) if isinstance(row[7], (int, float)) and row[7] else 0
        codigo = str(row[4]).strip()
        desc   = str(row[5]).strip() if row[5] else ''
        bd[codigo] = {
            'codigo': codigo, 'fuente': 'Gobernación de Boyacá',
            'capitulo': str(row[1]).strip() if row[1] else '',
            'subcapitulo': str(row[3]).strip() if row[3] else '',
            'descripcion': desc, 'unidad': str(row[6]).strip() if row[6] else '',
            'vr_mo': mo, 'vr_mat_equip': total - mo,
            'total': total, 'desc_norm': _norm(desc),
        }
    return bd


def _cargar_invias(ws):
    """
    Carga APUs INVIAS desde Excel descargado de hermes.invias.gov.co
    Columnas esperadas: ITEM, DESCRIPCION_ACTIVIDAD, UNIDAD,
    SUBTOTAL_MATERIALES, SUBTOTAL_EQUIPOS, SUBTOTAL_MANO_OBRA,
    SUBTOTAL_TRANSPORTE, COSTO_DIRECTO, NOMBRE_PROVINCIA
    """
    primera = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    cols = [str(v).upper().strip() if v else '' for v in primera]

    def _ci(nombre):  # column index
        for j, c in enumerate(cols):
            if nombre in c:
                return j
        return None

    col_item = _ci('ITEM')
    col_desc = _ci('DESCRIPCION_ACTIVIDAD') or _ci('DESCRIPCION')
    col_und  = _ci('UNIDAD')
    col_mat  = _ci('SUBTOTAL_MATERIALES') or _ci('MATERIALES')
    col_mo   = _ci('SUBTOTAL_MANO_OBRA') or _ci('MANO_OBRA') or _ci('MANO DE OBRA')
    col_eq   = _ci('SUBTOTAL_EQUIPOS') or _ci('EQUIPOS')
    col_tra  = _ci('SUBTOTAL_TRANSPORTE') or _ci('TRANSPORTE')
    col_tot  = _ci('COSTO_DIRECTO') or _ci('COSTO DIRECTO') or _ci('TOTAL')
    col_prov = _ci('NOMBRE_PROVINCIA') or _ci('PROVINCIA')

    if col_desc is None or col_tot is None:
        return {}

    bd = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row)
        total = row[col_tot] if col_tot < len(row) else None
        if not isinstance(total, (int, float)) or total <= 0:
            continue
        desc = str(row[col_desc]).strip() if col_desc < len(row) and row[col_desc] else ''
        if not desc:
            continue
        codigo = str(row[col_item]).strip() if col_item is not None and col_item < len(row) and row[col_item] else f'INV-{len(bd)+1}'
        und    = str(row[col_und]).strip() if col_und is not None and col_und < len(row) and row[col_und] else ''
        prov   = str(row[col_prov]).strip() if col_prov is not None and col_prov < len(row) and row[col_prov] else ''
        mat    = float(row[col_mat]) if col_mat is not None and col_mat < len(row) and isinstance(row[col_mat], (int,float)) else 0
        mo     = float(row[col_mo])  if col_mo  is not None and col_mo  < len(row) and isinstance(row[col_mo],  (int,float)) else 0
        eq     = float(row[col_eq])  if col_eq  is not None and col_eq  < len(row) and isinstance(row[col_eq],  (int,float)) else 0
        tra    = float(row[col_tra]) if col_tra  is not None and col_tra < len(row) and isinstance(row[col_tra], (int,float)) else 0

        bd[codigo] = {
            'codigo': codigo, 'fuente': f'INVIAS 2025-2{" · " + prov if prov else ""}',
            'descripcion': desc, 'unidad': und,
            'vr_mo': mo, 'vr_mat_equip': mat + eq,
            'subtotal_mat': mat, 'subtotal_mo': mo,
            'subtotal_eq': eq, 'subtotal_tra': tra,
            'total': float(total), 'desc_norm': _norm(desc),
            'provincia': prov,
            # Componentes desagregados para APU completo
            'componentes': _componentes_desde_subtotales(mat, mo, eq, tra, prov),
        }
    return bd


def _componentes_desde_subtotales(mat, mo, eq, tra, provincia=''):
    """Construye secciones mínimas desde subtotales INVIAS."""
    fuente = f'INVIAS 2025-2{" · " + provincia if provincia else ""}'
    secciones = {'materiales': [], 'herramientas': [], 'transporte': [], 'mano_de_obra': []}
    if mat > 0:
        secciones['materiales'].append({
            'description': 'Materiales (INVIAS referencia)',
            'unit': 'Glb', 'rend': 1.0, 'unit_price': mat,
            'parcial': mat, '_rend_ajustado': False, '_fuente': fuente,
        })
    if mo > 0:
        secciones['mano_de_obra'].append({
            'description': 'Mano de obra (INVIAS referencia)',
            'unit': 'Glb', 'rend': 1.0, 'unit_price': mo,
            'parcial': mo, '_rend_ajustado': False, '_fuente': fuente,
        })
    if eq > 0:
        secciones['herramientas'].append({
            'description': 'Equipos (INVIAS referencia)',
            'unit': 'Glb', 'rend': 1.0, 'unit_price': eq,
            'parcial': eq, '_rend_ajustado': False, '_fuente': fuente,
        })
    if tra > 0:
        secciones['transporte'].append({
            'description': 'Transporte (INVIAS referencia)',
            'unit': 'Glb', 'rend': 1.0, 'unit_price': tra,
            'parcial': tra, '_rend_ajustado': False, '_fuente': fuente,
        })
    return secciones


def _cargar_generico(ws):
    primera = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    cols = [str(v).upper().strip() if v else '' for v in primera]
    col_cod  = next((j for j, c in enumerate(cols) if 'CODIGO' in c or c == 'ITEM'), None)
    col_desc = next((j for j, c in enumerate(cols) if 'DESCRIPCION' in c or 'DESCRIPCI' in c), None)
    col_und  = next((j for j, c in enumerate(cols) if c in ('UNIDAD','UND','UN','UND.')), None)
    col_tot  = next((j for j, c in enumerate(cols)
                     if any(x in c for x in ('TOTAL','PRECIO UNIT','VALOR UNIT'))), None)
    if col_desc is None or col_tot is None:
        return {}
    bd = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row)
        total = row[col_tot] if col_tot < len(row) else None
        if not isinstance(total, (int, float)) or total <= 0: continue
        desc = str(row[col_desc]).strip() if col_desc < len(row) and row[col_desc] else ''
        if not desc: continue
        codigo = str(row[col_cod]).strip() if col_cod is not None and col_cod < len(row) and row[col_cod] else f'EXT-{len(bd)+1}'
        und = str(row[col_und]).strip() if col_und is not None and col_und < len(row) and row[col_und] else ''
        bd[codigo] = {
            'codigo': codigo, 'fuente': 'Base externa',
            'descripcion': desc, 'unidad': und,
            'vr_mo': 0, 'vr_mat_equip': float(total),
            'total': float(total), 'desc_norm': _norm(desc),
        }
    return bd


# ══════════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL DE CARGA — fuente genérica / INVIAS / Gobernación
# ══════════════════════════════════════════════════════════════════

def cargar_base_externa(archivo, nombre_fuente=None):
    try:
        wb = load_workbook(archivo, data_only=True)
    except Exception as e:
        return {}, None, f"No se pudo abrir el archivo: {e}"

    for nombre in wb.sheetnames:
        ws = wb[nombre]
        if ws.max_row < 5:
            continue
        fmt = _detectar_formato(ws)
        if fmt == 'boyaca':
            bd = _cargar_boyaca(ws)
            fuente = nombre_fuente or 'Gobernación de Boyacá'
        elif fmt == 'invias':
            bd = _cargar_invias(ws)
            fuente = nombre_fuente or 'INVIAS 2025-2'
        elif fmt == 'generico':
            bd = _cargar_generico(ws)
            fuente = nombre_fuente or 'Base externa'
        else:
            continue

        if bd:
            for v in bd.values():
                v['fuente'] = fuente
            return bd, fmt, None

    return {}, None, "No se reconoció el formato del archivo de base de precios."


# ══════════════════════════════════════════════════════════════════
# BÚSQUEDA POR SIMILITUD
# ══════════════════════════════════════════════════════════════════

def buscar_en_base(descripcion, unidad, bd, top_n=5, umbral_pct=30):
    palabras = {p for p in _norm(descripcion).split()
                if len(p) >= 4 and p not in _STOPWORDS}
    if not palabras:
        return []
    und_norm = _norm(unidad) if unidad else ''
    resultados = []
    for item in bd.values():
        palabras_item = set(item['desc_norm'].split())
        coincidencias = len(palabras & palabras_item)
        if coincidencias == 0:
            continue
        pct = coincidencias / len(palabras) * 100
        if pct < umbral_pct:
            continue
        bonus = 0.5 if und_norm and und_norm == _norm(item['unidad']) else 0
        resultados.append((coincidencias + bonus, pct, item))
    resultados.sort(key=lambda x: -x[0])
    return resultados[:top_n]


def construir_apu_desde_base(item_bd, valor_ofrecido):
    """
    Construye APU compatible con el generador desde un ítem de base externa.
    Usa componentes desagregados si existen (INVIAS), o subtotales si no.
    """
    total_ref = item_bd['total']

    # Si tiene componentes desagregados (formato INVIAS con subtotales)
    if 'componentes' in item_bd:
        secciones = item_bd['componentes']
        # Clonar para no mutar el original
        import copy
        secciones = copy.deepcopy(secciones)
    else:
        # Formato simple: dos líneas (MO + materiales/equipos)
        factor = valor_ofrecido / total_ref if total_ref > 0 else 1.0
        mo_ref = item_bd.get('vr_mo', 0)
        mat_eq_ref = item_bd.get('vr_mat_equip', total_ref)
        mo_esc = round(mo_ref * factor, 2)
        mat_esc = round(mat_eq_ref * factor, 2)
        diff = valor_ofrecido - mo_esc - mat_esc
        if mo_esc > 0:
            mo_esc = round(mo_esc + diff, 2)
        else:
            mat_esc = round(mat_esc + diff, 2)
        secciones = {
            'materiales':   [{'description': 'Materiales, insumos y equipos', 'unit': 'GL',
                               'rend': 1.0, 'unit_price': mat_esc, 'parcial': mat_esc,
                               '_rend_ajustado': False}] if mat_esc > 0 else [],
            'herramientas': [],
            'transporte':   [],
            'mano_de_obra': [{'description': 'Mano de obra (cuadrilla)', 'unit': 'HR',
                               'rend': 1.0, 'unit_price': mo_esc, 'parcial': mo_esc,
                               '_rend_ajustado': False}] if mo_esc > 0 else [],
        }

    return {
        **secciones,
        'fuente_bd':        item_bd['fuente'],
        'codigo_bd':        item_bd['codigo'],
        'total_referencia': total_ref,
    }


# ══════════════════════════════════════════════════════════════════
# CATÁLOGO POLICÍA NACIONAL 2026
# Fuente: 10. PRESUPUESTO APU ESPECIFICACIONES TECNICAS PARTICULARES.xlsx
# Código: 1LF-FR-0206 · Vigencia 2026
# ══════════════════════════════════════════════════════════════════

def cargar_catalogo_policia(ruta_archivo, zona: str = "A1"):
    try:
        xls = pd.ExcelFile(ruta_archivo)
    except Exception as e:
        return {}, {}, {}, {}, f"No se pudo abrir el catálogo Policía: {e}"

    zona = zona.upper()
    zona_idx = {"A1": 4, "A2": 5, "A3": 6, "A4": 7, "A5": 8, "A6": 9}
    col_z = zona_idx.get(zona, 4)

    actividades = {}
    try:
        df_act = pd.read_excel(ruta_archivo, sheet_name="LISTADO ACTIVIDADES", header=None)
        for _, row in df_act.iterrows():
            cod   = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
            desc  = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ""
            und   = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ""
            precio = row.iloc[col_z] if col_z < len(row) else None
            if not cod or not desc or cod in ("NaN", "nan", "CAP.", "ACTIVIDAD"):
                continue
            if not isinstance(precio, (int, float)) or precio <= 0:
                continue
            actividades[cod] = {
                "codigo": cod, "descripcion": desc, "unidad": und,
                "precio": float(precio), "zona": zona,
                "fuente": "Catálogo Policía 1LF-FR-0206 · 2026",
            }
    except Exception:
        pass

    insumos = {}
    try:
        df_ins = pd.read_excel(ruta_archivo, sheet_name="INSUMOS", header=0)
        for _, row in df_ins.iterrows():
            desc  = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            und   = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
            precio = row.iloc[2] if len(row) > 2 else None
            if not desc or desc in ("Descripción",): continue
            if not isinstance(precio, (int, float)) or precio <= 0: continue
            insumos[_norm(desc)] = {
                "descripcion": desc, "unidad": und, "precio": float(precio),
                "fuente": "Catálogo Policía 1LF-FR-0206 · 2026 · Insumos",
            }
    except Exception:
        pass

    equipos = {}
    try:
        df_eq = pd.read_excel(ruta_archivo, sheet_name="EQUIPOS", header=None)
        header_row = list(df_eq.iloc[1])
        col_zona_eq = None
        for j, h in enumerate(header_row):
            if h and isinstance(h, str) and zona in str(h).upper():
                col_zona_eq = j
                break
        if col_zona_eq is None:
            col_zona_eq = {"A1":2,"A2":3,"A3":4,"A4":5,"A5":6,"A6":7}.get(zona, 2)
        for _, row in df_eq.iloc[2:].iterrows():
            desc  = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            und   = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
            precio = row.iloc[col_zona_eq] if col_zona_eq < len(row) else None
            if not desc or not isinstance(precio, (int, float)) or precio <= 0: continue
            equipos[_norm(desc)] = {
                "descripcion": desc, "unidad": und, "precio_hora": float(precio),
                "zona": zona, "fuente": "Catálogo Policía 1LF-FR-0206 · 2026 · Equipos",
            }
    except Exception:
        pass

    mano_obra = {}
    try:
        df_mo = pd.read_excel(ruta_archivo, sheet_name="MANO DE OBRA", header=None)
        from salarios import MULTIPLICADORES_ZONA
        for _, row in df_mo.iterrows():
            desc = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            und  = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
            tarifa_base = row.iloc[2] if len(row) > 2 else None
            if not desc or len(desc) < 5: continue
            if not isinstance(tarifa_base, (int, float)) or tarifa_base <= 0: continue
            if any(x in desc.upper() for x in ("DESCRIPCI","SALARIO","CLASIFIC","PERSONAL","ESTRUCTURA")):
                if tarifa_base > 50_000_000: continue
            mult = MULTIPLICADORES_ZONA.get(zona, 1.0)
            mano_obra[_norm(desc)] = {
                "descripcion": desc, "unidad": und if und else "mes",
                "precio_hora": round(float(tarifa_base) * mult, 6),
                "precio_base": float(tarifa_base), "zona": zona,
                "fuente": "Catálogo Policía 1LF-FR-0206 · 2026 · Mano de Obra",
            }
    except Exception:
        pass

    return actividades, insumos, equipos, mano_obra, None


def buscar_actividad_policia(descripcion, unidad, actividades, top_n=5, umbral_pct=25):
    palabras = {p for p in _norm(descripcion).split()
                if len(p) >= 4 and p not in _STOPWORDS}
    if not palabras:
        return []
    und_norm = _norm(unidad) if unidad else ""
    resultados = []
    for item in actividades.values():
        palabras_item = set(_norm(item["descripcion"]).split())
        coincidencias = len(palabras & palabras_item)
        if coincidencias == 0: continue
        pct = coincidencias / len(palabras) * 100
        if pct < umbral_pct: continue
        bonus = 0.5 if und_norm and und_norm == _norm(item["unidad"]) else 0
        resultados.append((coincidencias + bonus, pct, item))
    resultados.sort(key=lambda x: -x[0])
    return resultados[:top_n]


# ══════════════════════════════════════════════════════════════════
# CASCADA DE FUENTES — función central del Camino B
# ══════════════════════════════════════════════════════════════════

def buscar_en_cascada(descripcion, unidad, valor_ofrecido, fuentes: list[dict]) -> dict | None:
    """
    Busca la actividad en múltiples fuentes en orden de prioridad.
    fuentes: lista de dicts con keys:
      - tipo: 'policia' | 'invias' | 'generica'
      - nombre: nombre para mostrar
      - bd o actividades: el diccionario de datos cargado
    Retorna dict compatible con el generador o None si no encontró nada.
    """
    for fuente in fuentes:
        tipo   = fuente.get('tipo', 'generica')
        nombre = fuente.get('nombre', 'Fuente externa')

        if tipo == 'policia':
            actividades = fuente.get('actividades', {})
            if not actividades:
                continue
            resultados = buscar_actividad_policia(descripcion, unidad, actividades, top_n=1)
            if not resultados:
                continue
            _, pct, item = resultados[0]
            # Construir APU desde actividad Policía (precio total, sin desagregar)
            precio_ref = item['precio']
            factor = valor_ofrecido / precio_ref if precio_ref > 0 else 1.0
            return {
                'fuente_bd':        f"{nombre} · {pct:.0f}% similitud",
                'codigo_bd':        item['codigo'],
                'total_referencia': precio_ref,
                'materiales':       [{'description': 'Materiales e insumos (Policía ref.)',
                                      'unit': 'Glb', 'rend': 1.0,
                                      'unit_price': round(precio_ref * factor * 0.60, 0),
                                      'parcial':    round(precio_ref * factor * 0.60, 0),
                                      '_rend_ajustado': False}],
                'herramientas':     [],
                'transporte':       [],
                'mano_de_obra':     [{'description': 'Mano de obra (Policía ref.)',
                                      'unit': 'Glb', 'rend': 1.0,
                                      'unit_price': round(precio_ref * factor * 0.40, 0),
                                      'parcial':    round(precio_ref * factor * 0.40, 0),
                                      '_rend_ajustado': False}],
            }

        else:  # 'invias' o 'generica'
            bd = fuente.get('bd', {})
            if not bd:
                continue
            resultados = buscar_en_base(descripcion, unidad, bd, top_n=1)
            if not resultados:
                continue
            _, pct, item_bd = resultados[0]
            item_bd = dict(item_bd)
            item_bd['fuente'] = f"{item_bd.get('fuente', nombre)} · {pct:.0f}% similitud"
            return construir_apu_desde_base(item_bd, valor_ofrecido)

    return None  # no encontrado en ninguna fuente


def generar_reporte_cobertura(items_con_apu, items_sin_apu) -> dict:
    """
    Genera reporte de cobertura para mostrar en la interfaz.
    Retorna dict con estadísticas y listas clasificadas.
    """
    total = len(items_con_apu) + len(items_sin_apu)

    # Clasificar los que tienen APU por fuente
    por_fuente = {}
    for item in items_con_apu:
        fuente = item.get('fuente_bd', 'Entidad contratante')
        if not fuente:
            fuente = 'Entidad contratante'
        por_fuente[fuente] = por_fuente.get(fuente, 0) + 1

    return {
        'total':          total,
        'con_apu':        len(items_con_apu),
        'sin_apu':        len(items_sin_apu),
        'pct_cobertura':  round(len(items_con_apu) / total * 100, 1) if total > 0 else 0,
        'por_fuente':     por_fuente,
        'items_sin_apu':  [{'code': i['code'], 'description': i['description'],
                             'unit': i['unit'], 'valor': i['valor_ofrecido']}
                            for i in items_sin_apu],
    }
