"""
Generador de APUs - Gerencia Legal Integral Colombia S.A.S.
v7: detección robusta — acepta cualquier variante de encabezado y número de ítem.
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io, re


# ══════════════════════════════════════════════════════════════════
# UTILIDADES DE DETECCIÓN — importadas desde detector.py
# ══════════════════════════════════════════════════════════════════
from numero_letras import numero_a_letras
from bases_externas import buscar_en_base, construir_apu_desde_base
from detector import (
    detectar_columnas as _detectar_columnas_ext,
    es_fila_item_valida as _es_fila_item_valida_ext,
    extraer_codigo as _extraer_codigo_ext,
    es_codigo_item,
    encontrar_hoja_presupuesto as _encontrar_hoja_presupuesto,
)

# Alias para compatibilidad con el resto del código
def _detectar_columnas(ws):     return _detectar_columnas_ext(ws)
def _es_fila_item(r, m):        return _es_fila_item_valida_ext(r, m)
def _extraer_codigo(r, m):      return _extraer_codigo_ext(r, m)

# ══════════════════════════════════════════════════════════════════
# LECTURA DE APUs EN HOJAS INDIVIDUALES (formato "APU X.XX")
# ══════════════════════════════════════════════════════════════════

def _leer_apu_hoja_individual(ws):
    """
    Lee un APU desde una hoja individual tipo SENA/Gobernación.
    Estructura: INSUMOS/MATERIALES, EQUIPO/HERRAMIENTA, MANO DE OBRA
    con columnas: ÍTEM | DESCRIPCIÓN | UNIDAD | V. UNITARIO | CANTIDAD | V. PARCIAL
    Retorna dict {total_referencia, materiales, herramientas, transporte, mano_de_obra}.
    """
    apu = {'total_referencia': 0,
           'materiales': [], 'herramientas': [], 'transporte': [], 'mano_de_obra': []}

    seccion_actual = None
    COL_DESC  = None  # se detecta dinámicamente
    COL_VUNIT = None
    COL_CANT  = None
    COL_UND   = None  # columna de unidad de medida del componente

    for row in ws.iter_rows(min_row=1, values_only=True):
        vals = [v for v in row if v is not None]
        if not vals: continue

        primera = str(vals[0]).upper().strip()

        # Detectar inicio de sección
        if any(x in primera for x in ('INSUMO','MATERIAL')):
            seccion_actual = 'materiales'
            COL_DESC = COL_VUNIT = COL_CANT = COL_UND = None
            continue
        if any(x in primera for x in ('EQUIPO','HERRAMIENTA','TRANSPORTE')):
            seccion_actual = 'herramientas'
            COL_DESC = COL_VUNIT = COL_CANT = COL_UND = None
            continue
        if any(x in primera for x in ('MANO DE OB','MANO OB')):
            seccion_actual = 'mano_de_obra'
            COL_DESC = COL_VUNIT = COL_CANT = COL_UND = None
            continue

        # Detectar fila de encabezados de la sección
        row_list = list(row)
        row_upper = [str(v).upper().strip() if v else '' for v in row_list]
        if 'V. UNITARIO' in row_upper or 'V.UNITARIO' in row_upper or 'UNITARIO' in ' '.join(row_upper):
            for j, t in enumerate(row_upper):
                if 'DESCRIPCI' in t: COL_DESC = j
                if 'V. UNIT' in t or 'UNITARIO' in t: COL_VUNIT = j
                if 'CANTIDAD' in t or 'CANT' in t: COL_CANT = j
                if t in ('UNIDAD', 'UND', 'UN', 'UNID') or t.startswith('UNIDAD'): COL_UND = j
            continue

        # Detectar total costo directo
        if 'TOTAL COSTO' in primera or 'TOTAL DIRECTO' in primera:
            for v in row:
                if isinstance(v, (int, float)) and v > 0:
                    apu['total_referencia'] = float(v)
                    break
            continue

        # Fila de componente: debe tener número, descripción y valor
        if seccion_actual is None: continue
        if COL_VUNIT is None: continue

        # Detectar si es fila de datos: tiene descripción + V.UNITARIO + CANTIDAD válidos
        if COL_VUNIT is None or COL_CANT is None: continue
        vunit_check = row_list[COL_VUNIT] if COL_VUNIT < len(row_list) else None
        cant_check  = row_list[COL_CANT]  if COL_CANT  < len(row_list) else None
        if not isinstance(vunit_check, (int,float)) or vunit_check <= 0: continue
        if not isinstance(cant_check,  (int,float)) or cant_check  <= 0: continue

        desc  = str(row_list[COL_DESC]).strip()  if COL_DESC  is not None and COL_DESC  < len(row_list) and row_list[COL_DESC]  else ''
        vunit = row_list[COL_VUNIT] if COL_VUNIT is not None and COL_VUNIT < len(row_list) else None
        cant  = row_list[COL_CANT]  if COL_CANT  is not None and COL_CANT  < len(row_list) else None

        # Leer unidad: primero desde columna detectada en encabezado, luego buscar entre DESC y VUNIT
        und = ''
        if COL_UND is not None and COL_UND < len(row_list) and row_list[COL_UND]:
            und = str(row_list[COL_UND]).strip()
        elif COL_DESC is not None and COL_VUNIT is not None:
            for j in range(COL_DESC + 1, COL_VUNIT):
                v = row_list[j] if j < len(row_list) else None
                if v and isinstance(v, str) and len(v.strip()) <= 10:
                    und = v.strip()
                    break

        if not desc or not isinstance(vunit, (int, float)) or not isinstance(cant, (int, float)):
            continue
        if vunit <= 0 or cant <= 0: continue

        comp = {'description': desc, 'unit': und,
                'rend': float(cant), 'unit_price': float(vunit)}
        apu[seccion_actual].append(comp)

    return apu

def _leer_hojas_apu_individuales(wb):
    """
    Lee todas las hojas que empiezan con 'APU' en un archivo de referencia.
    Retorna dict {codigo → apu_dict}.
    """
    bd = {}
    for nombre in wb.sheetnames:
        nombre_limpio = nombre.strip()
        if not nombre_limpio.upper().startswith('APU'):
            continue
        # Extraer código del nombre: "APU 2.16" → "2.16"
        partes = nombre_limpio.split()
        if len(partes) >= 2:
            codigo_hoja = partes[-1].strip()
        else:
            continue

        ws = wb[nombre]
        apu = _leer_apu_hoja_individual(ws)

        # También intentar extraer código desde la celda F6 col A
        try:
            codigo_celda = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))[0][0]
            if codigo_celda and isinstance(codigo_celda, float):
                codigo_hoja = f"{codigo_celda:.2f}".rstrip('0').rstrip('.')
        except Exception:
            pass

        if apu['total_referencia'] > 0 or any(apu[s] for s in ('materiales','herramientas','mano_de_obra')):
            bd[codigo_hoja] = apu

    return bd


# ══════════════════════════════════════════════════════════════════
# LECTOR DE BASE DE DATOS APU (archivo de referencia)
# ══════════════════════════════════════════════════════════════════

def _leer_hoja_apu_columnar(ws_apu):
    """
    Lee hoja APU usando detección inteligente de columnas.
    Soporta cualquier formato de entidad colombiana — detecta columnas por nombre.
    """
    import unicodedata, re as _re

    # ── Sinónimos por campo (más específicos primero) ─────────────────────
    SINONIMOS = {
        'col_codins':  ['CODINS','COD INS','COD. INS','CODIGO INSUMO','COD INSUMO'],
        'col_tipo':    ['TIPO','CLASE','CATEGORIA','CATEGORÍA','ROL','TIPO INSUMO'],
        'col_parcial': ['VR. PARCIAL','VR PARCIAL','PARCIAL','VALOR PARCIAL','V. PARCIAL'],
        'col_total':   ['VR. TOTAL','VR TOTAL','COSTO DIRECTO','COSTO TOTAL','VALOR TOTAL','PRECIO TOTAL','V. TOTAL','TOTAL UNITARIO','CD'],
        'col_unit':    ['UNITARIO','VR. UNIT','VR UNIT','PRECIO UNIT','VALOR UNIT','P.U.','VALOR UNITARIO','COSTO UNITARIO','TARIFA'],
        'col_rend':    ['REND','RENDIMIENTO','CANT','CANTIDAD','CONSUMO','DESPERDICIO','FACTOR','COEFICIENTE'],
        'col_und':     ['UNIDAD','UND','UN','UND.','U/M','UNID','MEDIDA'],
        'col_desc':    ['INSUMO / DESCRIPCIÓN','INSUMO/DESCRIPCION','INSUMO / DESCRIPCION','DESCRIPCION','DESCRIPCIÓN','DETALLE','INSUMO','NOMBRE','ACTIVIDAD'],
        'col_codigo':  ['CODIGO ITEM','COD. ITEM','CODIGO','COD','ITEM','ÍTEM','NUM','N°','REF'],
    }
    TIPOS_ACTIVIDAD = {'ACTIVIDAD','ITEM','ÍTEM','APU','ANALISIS','ANÁLISIS'}
    TIPOS_SECCION = {
        'mano_de_obra': ['CUADRILLA','MANO DE OBRA','MO','PERSONAL','JORNAL','OPERARIO','OBRERO'],
        'herramientas': ['EQUIPO','EQUIPOS','HERRAMIENTA','HERRAMIENTAS','MAQUINARIA'],
        'transporte':   ['TRANSPORTE','FLETE','ACARREO','MOVILIZACION','MOVILIZACIÓN'],
        'materiales':   ['INSUMO','INSUMOS','MATERIAL','MATERIALES','ANALISIS BASICO',
                         'ANÁLISIS BÁSICO','ENSAYO','SUBCONTRATO','OTRO'],
    }

    def _n(s):
        t = str(s).upper().strip()
        t = ''.join(c for c in unicodedata.normalize('NFD',t) if unicodedata.category(c)!='Mn')
        return _re.sub(r'\s+',' ',t)

    def _score(header, syns):
        h = _n(header)
        for s in syns:
            sn = _n(s)
            if h == sn:      return 3
            if sn in h:      return 2
            if h in sn and len(h) >= 4: return 1
        return 0

    def _clasificar(tipo_val):
        t = _n(tipo_val)
        for sec, words in TIPOS_SECCION.items():
            for w in words:
                wn = _n(w)
                if wn == t: return sec
                if len(wn) >= 3 and _re.search(r''+_re.escape(wn)+r'', t):
                    return sec
        return 'materiales'

    def _es_actividad(tipo_val):
        t = _n(tipo_val)
        return any(_n(x) in t or t in _n(x) for x in TIPOS_ACTIVIDAD)

    # ── Detectar fila de encabezados ──────────────────────────────────────
    mejor_fila = 0; mejor_score = 0; mejor_mapa = {}; mejor_enc = []
    for fi in range(min(10, ws_apu.max_row)):
        fila = [c.value for c in ws_apu[fi+1]]
        if not any(fila): continue
        asignado = {}; mapa = {}; total_s = 0
        for campo, syns in SINONIMOS.items():
            col_scores = {ci: _score(str(cel),syns) for ci,cel in enumerate(fila)
                          if cel and _score(str(cel),syns) > 0 and ci not in asignado}
            if not col_scores: continue
            best_col = max(col_scores, key=col_scores.get)
            mapa[campo] = best_col
            asignado[best_col] = campo
            total_s += col_scores[best_col]
        if total_s > mejor_score:
            mejor_score=total_s; mejor_fila=fi; mejor_mapa=mapa
            mejor_enc=[str(c) if c else '' for c in fila]

    # Confianza mínima
    criticos = {'col_codigo','col_desc','col_rend','col_unit','col_total'}
    if len(criticos & set(mejor_mapa)) < 2:
        return {}   # formato no reconocido

    i_cod  = mejor_mapa.get('col_codigo')
    i_cins = mejor_mapa.get('col_codins')
    i_desc = mejor_mapa.get('col_desc')
    i_tipo = mejor_mapa.get('col_tipo')
    i_und  = mejor_mapa.get('col_und')
    i_rend = mejor_mapa.get('col_rend')
    i_unit = mejor_mapa.get('col_unit')
    i_parc = mejor_mapa.get('col_parcial')
    i_tot  = mejor_mapa.get('col_total')

    def _get(row, idx):
        return row[idx] if (idx is not None and idx < len(row)) else None

    # ── Detectar variante ─────────────────────────────────────────────────
    tipos_muestra = set()
    codins_muestra = set()
    for row in ws_apu.iter_rows(min_row=mejor_fila+2, max_row=mejor_fila+22, values_only=True):
        v = _get(row, i_tipo)
        if v: tipos_muestra.add(_n(str(v)))
        v2 = _get(row, i_cins)
        if v2: codins_muestra.add(str(v2).strip())

    tipos_conocidos = TIPOS_ACTIVIDAD | {_n(t) for lst in TIPOS_SECCION.values() for t in lst}
    if tipos_muestra & tipos_conocidos:
        variante = 'tipo_columna'
    elif '-' in codins_muestra:
        variante = 'codins_marca'
    else:
        variante = 'desconocido'

    # ── Leer filas ────────────────────────────────────────────────────────
    bd = {}; current = None
    for row in ws_apu.iter_rows(min_row=mejor_fila+2, values_only=True):
        if not any(row): continue
        cod_v  = _get(row, i_cod)
        codins = str(_get(row, i_cins) or '').strip()
        desc   = str(_get(row, i_desc) or '').strip()
        tipo_v = str(_get(row, i_tipo) or '').strip()
        und    = str(_get(row, i_und)  or '').strip()
        rend   = _get(row, i_rend)
        unit   = _get(row, i_unit)
        parc   = _get(row, i_parc)
        total  = _get(row, i_tot)
        cod_s  = str(cod_v).strip() if cod_v else ''

        es_act = False
        if variante == 'tipo_columna':
            es_act = bool(tipo_v) and _es_actividad(tipo_v) and codins == '-'
        elif variante == 'codins_marca':
            es_act = codins == '-'
        else:
            es_act = (isinstance(total,(int,float)) and total > 0
                      and not isinstance(rend,(int,float))
                      and not isinstance(unit,(int,float)) and bool(cod_s))

        if es_act and cod_s:
            current = cod_s
            bd[current] = {
                'total_referencia': float(total) if isinstance(total,(int,float)) else 0,
                'descripcion': desc,
                'materiales':[],'herramientas':[],'transporte':[],'mano_de_obra':[],
            }
            continue

        if not current: continue
        if not isinstance(rend,(int,float)) or not isinstance(unit,(int,float)): continue
        if rend == 0 and unit == 0: continue

        parc_v = float(parc) if isinstance(parc,(int,float)) else round(float(rend)*float(unit),0)
        comp = {'description':desc,'unit':und,'rend':float(rend),'unit_price':float(unit),
                'parcial':parc_v,'_rend_ajustado':False}

        if tipo_v:
            sec = _clasificar(tipo_v)
        else:
            d = _n(desc)
            if any(x in d for x in ('CUADRILLA','JORNAL','OFICIAL','AYUDANTE')):
                sec = 'mano_de_obra'
            elif any(x in d for x in ('EQUIPO','MAQUINARIA','HERRAMIENTA')):
                sec = 'herramientas'
            elif any(x in d for x in ('TRANSPORTE','FLETE','ACARREO')):
                sec = 'transporte'
            else:
                sec = 'materiales'
        bd[current][sec].append(comp)

    return bd

def _construir_mapa_codigos(wb):
    """
    Lee la hoja PRESUPUESTO (o similar) del archivo de referencia
    y construye un mapeo {cod_gobernacion → item_num_interno}.
    Sirve para cruzar la oferta (usa cod gober) con los APUs (usan num interno).
    """
    mapa = {}  # cod_gober → item_num  (ej: "2.01.34" → "2.01")
    ws = None
    for nombre in wb.sheetnames:
        if 'PRESUPUESTO' in nombre.upper() or 'RESUMEN' in nombre.upper():
            ws = wb[nombre]; break
    if not ws: return mapa

    _, col_mapa = _detectar_columnas(ws)
    # Buscar las dos primeras columnas numéricas
    # En CIMM: col A = ítem interno, col B = código gobernación
    for row in ws.iter_rows(min_row=1, max_row=200, values_only=True):
        item_num = row[0]
        cod_gob  = row[1]
        if (item_num and isinstance(item_num, float) and item_num > 0
                and cod_gob and isinstance(cod_gob, str) and '.' in cod_gob):
            num_str = f"{item_num:.2f}".rstrip('0').rstrip('.')
            mapa[cod_gob.strip()] = num_str
    return mapa


def _leer_apu_formato_presupuesto_directo(ws):
    """
    Lee APUs del formato de presupuesto interno (Pascual Bravo / similar).
    Todos los APUs en una sola hoja con patrón:
      Señal: 'ANALISIS DE PRECIOS UNITARIOS - APU'
      Encabezado: col0='ITEM' col5='UNIDAD' col6='COSTO DIRECTO'
      Ítem:       col0=código, col1=desc, col5=unidad, col6=costo_directo
      Secciones:  col0=10000/20000/30000/40000/50000
      Componente: col0=cod_num, col1=desc, col2=und, col3=cant, col4=rend_o_desperd, col5=vr_unit, col6=vr_total
      Fin APU:    col1='TOTAL UNITARIO'
    """
    SECCIONES = {
        10000: 'materiales', 20000: 'mano_de_obra',
        30000: 'herramientas', 40000: 'materiales', 50000: 'transporte',
    }
    COD_SEC = set(SECCIONES.keys())

    apus = {}
    current_apu = None
    estado = 'buscando'
    seccion = None
    es_mo = False

    for row in ws.iter_rows(values_only=True):
        c = list(row) + [None] * 11
        c0,c1,c2,c3,c4,c5,c6 = c[0],c[1],c[2],c[3],c[4],c[5],c[6]

        if c0 and str(c0).strip() == 'ANALISIS DE PRECIOS UNITARIOS - APU':
            estado = 'enc_item'; continue

        if estado == 'enc_item' and c0 and str(c0).strip() == 'ITEM':
            estado = 'item'; continue

        if estado == 'item':
            estado = 'leyendo'; seccion = None
            if c0 and c1 and c6 is not None and isinstance(c6,(int,float)) and c6 > 0:
                code = str(c0).strip()
                current_apu = {
                    'code': code, 'description': str(c1).strip(),
                    'unit': str(c5).strip() if c5 else '',
                    'total_referencia': float(c6),
                    'materiales':[],'herramientas':[],'transporte':[],'mano_de_obra':[],
                }
                apus[code] = current_apu
            else:
                current_apu = None
            continue

        if estado != 'leyendo' or current_apu is None: continue

        if c1 and 'TOTAL UNITARIO' in str(c1).upper():
            current_apu = None; seccion = None; estado = 'buscando'; continue

        if isinstance(c0, int) and c0 in COD_SEC:
            seccion = SECCIONES[c0]; es_mo = (c0 == 20000); continue

        if c0 and str(c0).strip() in ('COD','#N/A'): continue
        if seccion is None: continue

        if (isinstance(c0, int) and c0 not in COD_SEC and
                c1 and isinstance(c1, str) and len(c1.strip()) > 2 and
                c5 is not None and isinstance(c5,(int,float)) and c5 > 0 and
                c6 is not None and isinstance(c6,(int,float)) and c6 > 0):
            # Derivar rendimiento desde VR.TOTAL / VR.UNITARIO
            # Esto captura desperdicio (materiales), CANT2 (equipos) y jornales/rendimiento (MO)
            vr_total = float(c6)
            vr_unit  = float(c5)
            rend = vr_total / vr_unit if vr_unit > 0 else 0
            if rend > 0:
                current_apu[seccion].append({
                    'description': str(c1).strip(), 'unit': str(c2).strip() if c2 else '',
                    'rend': round(rend, 6), 'unit_price': vr_unit,
                    'parcial': vr_total,
                })
    return apus

def leer_base_datos_apu(archivo_referencia):
    """
    Lee un archivo de referencia e intenta extraer APUs por cualquier método.
    Soporta: hoja A.P.U columnar, hojas individuales "APU X.XX", y PRESUPUESTO.
    """
    try:
        if hasattr(archivo_referencia, 'read'):
            data = archivo_referencia.read()
            archivo_referencia = io.BytesIO(data)
        wb = load_workbook(archivo_referencia, data_only=True)
    except Exception as e:
        return {}, f"No se pudo abrir el archivo de referencia: {e}"

    bd = {}

    # Método 1: hojas individuales "APU X.XX"
    bd_ind = _leer_hojas_apu_individuales(wb)
    if bd_ind:
        bd.update(bd_ind)

    # Método 2: hoja columnar A.P.U / APU
    for nombre in wb.sheetnames:
        if 'A.P.U' in nombre.upper() or nombre.upper() == 'APU':
            bd_col = _leer_hoja_apu_columnar(wb[nombre])
            bd.update(bd_col)
            break

    # Método 3: formato presupuesto directo (hoja con múltiples APUs, señal textual)
    # Buscar hojas con nombre APU, APU Obra Civil, APU Extraccion, APUS ELEC, etc.
    for nombre in wb.sheetnames:
        n_up = nombre.upper().strip()
        if n_up.startswith('APU') or 'APU' in n_up:
            bd_pd = _leer_apu_formato_presupuesto_directo(wb[nombre])
            if bd_pd:
                bd.update(bd_pd)

    # Método 4: construir puente entre código gobernación y número interno
    # (necesario cuando el archivo de referencia tiene APUs por número interno
    #  pero la oferta usa códigos de gobernación)
    mapa_codigos = _construir_mapa_codigos(wb)
    for cod_gob, num_int in mapa_codigos.items():
        if num_int in bd and cod_gob not in bd:
            bd[cod_gob] = bd[num_int]  # registrar bajo ambas claves

    return bd, None if bd else "No se encontraron APUs en el archivo de referencia."


# ══════════════════════════════════════════════════════════════════
# LECTURA PRINCIPAL DEL PROCESO
# ══════════════════════════════════════════════════════════════════

def leer_oferta_economica(uploaded_file, bd_referencia=None):
    try:
        data = uploaded_file.read()
        wb = load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:
        return None, f"No se pudo abrir el archivo: {e}"

    ws_pres = _encontrar_hoja_presupuesto(wb)
    fila_enc, mapa = _detectar_columnas(ws_pres)

    if fila_enc is None or 'col_valor' not in mapa:
        return None, (
            f"No se encontró columna de valor unitario en la hoja '{ws_pres.title}'.\n"
            f"Hojas disponibles: {', '.join(wb.sheetnames)}\n\n"
            f"Se buscaron variantes de: VALOR UNITARIO, Vr. Unit., V. Unitario, P.U., etc."
        )

    col_valor = mapa['col_valor']
    col_desc  = mapa.get('col_desc')
    col_und   = mapa.get('col_und')

    items_proceso = {}
    for row in ws_pres.iter_rows(min_row=fila_enc + 2, values_only=True):
        if not _es_fila_item(row, mapa): continue
        codigo = _extraer_codigo(row, mapa)
        desc   = str(row[col_desc]).strip() if (col_desc is not None and col_desc < len(row) and row[col_desc]) else codigo
        und    = str(row[col_und]).strip()  if (col_und  is not None and col_und  < len(row) and row[col_und])  else ''
        valor  = float(row[col_valor])
        if codigo not in items_proceso:
            items_proceso[codigo] = {
                'code': codigo, 'description': desc, 'unit': und,
                'valor_ofrecido': valor,
                'materiales': [], 'herramientas': [],
                'transporte': [], 'mano_de_obra': [],
                'tiene_apu': False,
            }

    if not items_proceso:
        # Segundo intento: buscar otra hoja en el archivo
        hojas_restantes = [h for h in wb.sheetnames if h != ws_pres.title]
        for hoja_alt in hojas_restantes:
            ws_alt = wb[hoja_alt]
            fila_alt, mapa_alt = _detectar_columnas(ws_alt)
            if fila_alt is None or 'col_valor' not in mapa_alt:
                continue
            items_alt = {}
            col_v2 = mapa_alt['col_valor']
            col_d2 = mapa_alt.get('col_desc')
            col_u2 = mapa_alt.get('col_und')
            for row in ws_alt.iter_rows(min_row=fila_alt + 2, values_only=True):
                if not _es_fila_item(row, mapa_alt): continue
                cod2 = _extraer_codigo(row, mapa_alt)
                desc2 = str(row[col_d2]).strip() if (col_d2 is not None and col_d2 < len(row) and row[col_d2]) else cod2
                und2  = str(row[col_u2]).strip()  if (col_u2 is not None and col_u2 < len(row) and row[col_u2])  else ''
                val2  = float(row[col_v2])
                if cod2 not in items_alt:
                    items_alt[cod2] = {'code':cod2,'description':desc2,'unit':und2,
                                       'valor_ofrecido':val2,'materiales':[],'herramientas':[],
                                       'transporte':[],'mano_de_obra':[],'tiene_apu':False}
            if items_alt:
                items_proceso = items_alt
                ws_pres = ws_alt
                mapa = mapa_alt
                break

    if not items_proceso:
        return None, (
            f"No se encontraron ítems con valor unitario en '{ws_pres.title}'.\n"
            f"Encabezados detectados en fila {fila_enc+1}: {mapa}"
        )

    # ── APU del mismo archivo (columnar o individual) ─────────────────────────
    bd_interna = {}
    bd_ind = _leer_hojas_apu_individuales(wb)
    if bd_ind:
        bd_interna.update(bd_ind)
    for nombre in wb.sheetnames:
        if 'A.P.U' in nombre.upper() or nombre.upper() == 'APU':
            bd_interna.update(_leer_hoja_apu_columnar(wb[nombre]))
            break

    # ── Combinar fuentes ──────────────────────────────────────────────────────
    bd_ext = bd_referencia or {}
    bd_combinada = {**bd_ext, **bd_interna}  # interna tiene prioridad

    # ── Cruzar y escalar con reglas de prioridad ──────────────────────────────
    for code, item in items_proceso.items():
        apu = bd_combinada.get(code)
        if not apu:
            continue
        ref = apu['total_referencia']
        if ref <= 0:
            continue

        ofrecido = item['valor_ofrecido']

        # Copiar componentes de referencia — precios unitarios intactos (tarifas oficiales)
        for sec in ('materiales', 'herramientas', 'transporte', 'mano_de_obra'):
            item[sec] = [dict(c) for c in apu[sec]]
            for c in item[sec]:
                c['_rend_ajustado'] = False   # marca de trazabilidad

        def _total_sec(comps):
            return sum(c['rend'] * c['unit_price'] for c in comps
                       if c['unit_price'] > 0 and c['rend'] > 0)

        mo_ref    = _total_sec(item['mano_de_obra'])
        equip_ref = _total_sec(item['herramientas']) + _total_sec(item['transporte'])
        mat_ref   = _total_sec(item['materiales'])
        total_ref = mo_ref + equip_ref + mat_ref

        # Si el precio ofrecido ya cuadra con la referencia (tolerancia $1), no tocar nada
        if abs(ofrecido - total_ref) <= 1:
            for sec in ('materiales', 'herramientas', 'transporte', 'mano_de_obra'):
                for c in item[sec]:
                    c['parcial'] = round(c['rend'] * c['unit_price'], 0)
            item['tiene_apu'] = True
            continue

        delta_restante = ofrecido - total_ref

        # ── PASO 1: ajustar RENDIMIENTO de mano de obra (columna F) ──────────
        # Regla: modificar solo el rendimiento, nunca el valor unitario (tarifa oficial).
        # Límites técnicos: mín 20% del original, máx 300% del original.
        if mo_ref > 0:
            mo_objetivo = mo_ref + delta_restante
            mo_min      = mo_ref * 0.20
            mo_max      = mo_ref * 3.00
            mo_ajustado = max(mo_min, min(mo_max, mo_objetivo))
            factor_mo   = mo_ajustado / mo_ref
            if abs(factor_mo - 1.0) > 0.0001:   # solo si hay cambio real
                for c in item['mano_de_obra']:
                    if c['unit_price'] > 0 and c['rend'] > 0:
                        c['rend']           = round(c['rend'] * factor_mo, 6)
                        c['_rend_ajustado'] = True
            delta_restante -= (mo_ajustado - mo_ref)

        # ── PASO 2: si queda diferencia, ajustar RENDIMIENTO de herramienta ──
        # Solo si el delta residual supera $1 (diferencia real, no redondeo).
        if equip_ref > 0 and abs(delta_restante) > 1:
            equip_objetivo = equip_ref + delta_restante
            equip_min      = equip_ref * 0.20
            equip_max      = equip_ref * 3.00
            equip_ajustado = max(equip_min, min(equip_max, equip_objetivo))
            factor_eq      = equip_ajustado / equip_ref
            if abs(factor_eq - 1.0) > 0.0001:
                for sec in ('herramientas', 'transporte'):
                    for c in item[sec]:
                        if c['unit_price'] > 0 and c['rend'] > 0:
                            c['rend']           = round(c['rend'] * factor_eq, 6)
                            c['_rend_ajustado'] = True
            delta_restante -= (equip_ajustado - equip_ref)

        # ── PASO 3 (último recurso): ajustar VALOR UNITARIO de insumos ───────
        # Solo si todavía queda diferencia > $1 después de MO y equipos.
        # Se ajusta el precio unitario (columna G), no el rendimiento.
        # Límites: mín 60% del original, máx 160% del original.
        if mat_ref > 0 and abs(delta_restante) > 1:
            mat_objetivo = mat_ref + delta_restante
            mat_min      = mat_ref * 0.60
            mat_max      = mat_ref * 1.60
            mat_ajustado = max(mat_min, min(mat_max, mat_objetivo))
            factor_mat   = mat_ajustado / mat_ref
            for c in item['materiales']:
                if c['unit_price'] > 0 and c['rend'] > 0:
                    c['unit_price']     = max(1, int(round(c['unit_price'] * factor_mat, 0)))
                    c['_rend_ajustado'] = True   # marca: este componente fue tocado

        # ── Recalcular parciales finales ──────────────────────────────────────
        for sec in ('materiales', 'herramientas', 'transporte', 'mano_de_obra'):
            for c in item[sec]:
                c['parcial'] = round(c['rend'] * c['unit_price'], 0)

        item['tiene_apu'] = True

        # ── Diagnóstico de mano de obra ───────────────────────────────────────
        mo_final   = sum(c['rend'] * c['unit_price'] for c in item['mano_de_obra'])
        total_comp = sum(
            c['rend'] * c['unit_price']
            for sec in ('materiales','herramientas','transporte','mano_de_obra')
            for c in item.get(sec, [])
        )
        pct_mo = (mo_final / total_comp * 100) if total_comp > 0 else 0
        item['alerta_mo'] = (pct_mo < 5 and len(item['mano_de_obra']) > 0)
        item['pct_mo']    = round(pct_mo, 1)

    con_apu = [i for i in items_proceso.values() if i['tiene_apu']]
    sin_apu = [i for i in items_proceso.values() if not i['tiene_apu']]

    return {
        'items_con_apu':  con_apu,
        'items_sin_apu':  sin_apu,
        'total_proceso':  len(items_proceso),
        'tiene_hoja_apu': bool(bd_interna),
        'hoja_usada':     ws_pres.title,
    }, None


# ══════════════════════════════════════════════════════════════════
# GENERACIÓN DEL EXCEL
# ══════════════════════════════════════════════════════════════════

def _safe_name(code):
    name = str(code)[:31]
    for c in '/\\?*:[]': name = name.replace(c, '-')
    return name

def _sort_key_codigo(code):
    """Ordena códigos tipo '2.01', '2.10', '08B.01', 'A.1.3' de forma numérica correcta."""
    partes = re.split(r'[.\-\s]', str(code).strip())
    resultado = []
    for p in partes:
        try:
            resultado.append((0, int(p)))
        except ValueError:
            try:
                resultado.append((0, float(p)))
            except ValueError:
                resultado.append((1, p.upper()))
    return resultado

def _build_apu_sheet(ws, item, include_aiu=False, aiu_pct=0.0):
    bold9  = Font(bold=True, name='Arial', size=9)
    norm9  = Font(bold=False, name='Arial', size=9)
    thin   = Side(style='thin')
    brd    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_w = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    right  = Alignment(horizontal='right',  vertical='center')
    gray   = PatternFill('solid', fgColor='D9D9D9')
    blue   = PatternFill('solid', fgColor='BDD7EE')
    CUR    = '$ #,##0'

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 4
    ws.column_dimensions['C'].width = 4
    ws.column_dimensions['D'].width = 4
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15

    azul_osc = PatternFill('solid', fgColor='1B3A6B')
    naranja   = PatternFill('solid', fgColor='FCE4D6')
    blanco_f  = Font(bold=True, name='Arial', size=9, color='FFFFFF')

    r = 1
    # ── Fila 1: ACTIVIDAD (descripción) — visible y prominente ──────────────────
    ws.row_dimensions[r].height = 38
    ws.merge_cells(f'A{r}:F{r}')
    ws[f'A{r}'] = item['description']
    ws[f'A{r}'].font = Font(bold=True, name='Arial', size=11)
    ws[f'A{r}'].alignment = left_w
    ws[f'A{r}'].border = brd
    ws[f'A{r}'].fill = blue

    ws[f'G{r}'] = 'ÍTEM'
    ws[f'G{r}'].font = bold9; ws[f'G{r}'].alignment = center
    ws[f'G{r}'].border = brd; ws[f'G{r}'].fill = gray

    ws[f'H{r}'] = item['code']
    ws[f'H{r}'].font = Font(bold=True, name='Arial', size=10)
    ws[f'H{r}'].alignment = center
    ws[f'H{r}'].border = brd
    r += 1

    # ── Fila 2: UNIDAD DE MEDIDA + alerta MO si aplica ─────────────────────────
    ws.row_dimensions[r].height = 18
    ws.merge_cells(f'A{r}:C{r}')
    ws[f'A{r}'] = 'UNIDAD DE MEDIDA'
    ws[f'A{r}'].font = bold9; ws[f'A{r}'].alignment = center
    ws[f'A{r}'].border = brd; ws[f'A{r}'].fill = gray

    ws.merge_cells(f'D{r}:F{r}')
    ws[f'D{r}'] = item['unit']
    ws[f'D{r}'].font = Font(bold=True, name='Arial', size=10)
    ws[f'D{r}'].alignment = center; ws[f'D{r}'].border = brd

    # Alerta de mano de obra baja
    alerta_mo = item.get('alerta_mo', False)
    pct_mo    = item.get('pct_mo', None)
    ws.merge_cells(f'G{r}:H{r}')
    if alerta_mo:
        aviso = f'⚠️ MO baja ({pct_mo}%)' if pct_mo is not None else '⚠️ MO baja'
        ws[f'G{r}'] = aviso
        ws[f'G{r}'].fill = naranja
    else:
        ws[f'G{r}'] = ''
    ws[f'G{r}'].font = bold9; ws[f'G{r}'].alignment = center; ws[f'G{r}'].border = brd
    r += 1

    def write_section(titulo, col_header, componentes):
        nonlocal r
        ws.row_dimensions[r].height = 14
        ws.merge_cells(f'A{r}:H{r}')
        ws[f'A{r}'] = titulo
        ws[f'A{r}'].font = bold9; ws[f'A{r}'].alignment = center
        ws[f'A{r}'].fill = gray;  ws[f'A{r}'].border = brd
        r += 1
        ws.row_dimensions[r].height = 14
        ws.merge_cells(f'A{r}:D{r}')
        for col, val in [('A','DESCRIPCION'),('E','UNIDAD'),
                         ('F',col_header),('G','VR UNIT'),('H','VR TOTAL')]:
            ws[f'{col}{r}'] = val
            ws[f'{col}{r}'].font = bold9
            ws[f'{col}{r}'].alignment = center
            ws[f'{col}{r}'].border = brd
        r += 1
        start = r
        n = max(len(componentes), 2)
        for i in range(n):
            ws.row_dimensions[r].height = 13
            comp = componentes[i] if i < len(componentes) else None
            ws.merge_cells(f'A{r}:D{r}')
            if comp:
                ws[f'A{r}'] = comp['description']
                ws[f'E{r}'] = comp['unit'] if comp.get('unit') else ''
                ws[f'F{r}'] = round(comp['rend'], 4)
                ws[f'G{r}'] = int(round(comp['unit_price'], 0))
                # Regla crítica: fila ajustada → =F*G sin redondeo; intacta → =ROUND(F*G,0)
                if comp.get('_rend_ajustado'):
                    ws[f'H{r}'] = f'=F{r}*G{r}'
                else:
                    ws[f'H{r}'] = f'=ROUND(F{r}*G{r},0)' 
            for col, al, fmt in [
                ('A',left_w,None),('E',center,None),
                ('F',right,'#,##0.0000'),('G',right,CUR),('H',right,CUR)]:
                ws[f'{col}{r}'].font = norm9
                ws[f'{col}{r}'].border = brd
                ws[f'{col}{r}'].alignment = al
                if fmt: ws[f'{col}{r}'].number_format = fmt
            r += 1
        end = r - 1
        ws.row_dimensions[r].height = 14
        ws.merge_cells(f'A{r}:G{r}')
        ws[f'A{r}'] = f'SUBTOTAL {titulo}'
        ws[f'A{r}'].font = bold9; ws[f'A{r}'].alignment = right
        ws[f'A{r}'].fill = gray;  ws[f'A{r}'].border = brd
        ws[f'H{r}'] = f'=SUM(H{start}:H{end})'
        ws[f'H{r}'].font = bold9; ws[f'H{r}'].alignment = right
        ws[f'H{r}'].fill = gray;  ws[f'H{r}'].border = brd
        ws[f'H{r}'].number_format = CUR
        sub = r; r += 1
        return sub

    row_mat = write_section('MATERIALES',               'CANTID.', item['materiales'])
    row_her = write_section('HERRAMIENTAS Y/O EQUIPOS', 'RENDIM.', item['herramientas'])
    row_tra = write_section('TRANSPORTE',               'CANTID.', item['transporte'])
    row_mdo = write_section('MANO DE OBRA',             'RENDIM.', item['mano_de_obra'])

    aiu_row = None
    if include_aiu and aiu_pct > 0:
        ws.row_dimensions[r].height = 14
        ws.merge_cells(f'A{r}:D{r}')
        ws[f'A{r}'] = 'A.I.U.'
        ws[f'A{r}'].font = bold9; ws[f'A{r}'].alignment = left_w; ws[f'A{r}'].border = brd
        ws[f'E{r}'] = '%'
        ws[f'E{r}'].font = bold9; ws[f'E{r}'].alignment = center; ws[f'E{r}'].border = brd
        ws[f'F{r}'] = aiu_pct
        ws[f'F{r}'].font = bold9; ws[f'F{r}'].alignment = right
        ws[f'F{r}'].border = brd; ws[f'F{r}'].number_format = '0.00%'
        ws[f'G{r}'].border = brd
        ws[f'H{r}'] = f'=ROUND((H{row_mat}+H{row_her}+H{row_tra}+H{row_mdo})*F{r},0)'
        ws[f'H{r}'].font = bold9; ws[f'H{r}'].alignment = right
        ws[f'H{r}'].border = brd; ws[f'H{r}'].number_format = CUR
        aiu_row = r; r += 1

    # ── PRECIO UNITARIO ─────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 18
    ws.merge_cells(f'A{r}:E{r}')
    ws[f'A{r}'] = 'PRECIO UNITARIO'
    ws[f'A{r}'].font = Font(bold=True, name='Arial', size=10)
    ws[f'A{r}'].alignment = center; ws[f'A{r}'].fill = blue; ws[f'A{r}'].border = brd
    ws.merge_cells(f'F{r}:H{r}')
    valor_precio = int(round(item['valor_ofrecido'], 0))
    ws[f'F{r}'] = valor_precio
    ws[f'F{r}'].font = Font(bold=True, name='Arial', size=10)
    ws[f'F{r}'].alignment = right; ws[f'F{r}'].fill = blue
    ws[f'F{r}'].border = brd; ws[f'F{r}'].number_format = CUR
    r += 1

    # ── TOTAL EN LETRAS ───────────────────────────────────────────────────
    gold = PatternFill('solid', fgColor='FFF2CC')
    ws.row_dimensions[r].height = 22
    ws.merge_cells(f'A{r}:H{r}')
    letras = numero_a_letras(valor_precio)
    ws[f'A{r}'] = letras
    ws[f'A{r}'].font = Font(bold=True, name='Arial', size=9, italic=True)
    ws[f'A{r}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws[f'A{r}'].fill = gold
    ws[f'A{r}'].border = brd


def _build_resumen_sheet(wb, items_generados):
    """
    Crea la hoja RESUMEN (primera hoja del workbook) con tabla de todos los APUs:
    Ítem | Descripción | Unidad | Precio ofrecido | Precio APU | Cierre | N° comp | Fuente
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(title='RESUMEN', index=0)

    # Estilos
    bold   = Font(bold=True, name='Arial', size=9)
    norm   = Font(bold=False, name='Arial', size=9)
    thin   = Side(style='thin')
    brd    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    right  = Alignment(horizontal='right',  vertical='center')
    azul   = PatternFill('solid', fgColor='1B3A6B')
    gris   = PatternFill('solid', fgColor='D9D9D9')
    verde  = PatternFill('solid', fgColor='E2EFDA')
    rojo   = PatternFill('solid', fgColor='FCE4D6')
    ambar  = PatternFill('solid', fgColor='FFF2CC')
    blanco = Font(bold=True, name='Arial', size=9, color='FFFFFF')
    CUR    = '$ #,##0'

    # Anchos de columna
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 14   # MO%
    ws.column_dimensions['I'].width = 35   # FUENTE

    # Título
    ws.row_dimensions[1].height = 28
    ws.merge_cells('A1:I1')
    ws['A1'] = 'RESUMEN DE APUs GENERADOS — Gerencia Legal Integral Colombia S.A.S.'
    ws['A1'].font = blanco
    ws['A1'].fill = azul
    ws['A1'].alignment = center

    # Encabezados
    ws.row_dimensions[2].height = 30
    encabezados = [
        ('A2', 'ÍTEM'),
        ('B2', 'DESCRIPCIÓN'),
        ('C2', 'UND.'),
        ('D2', 'PRECIO OFRECIDO'),
        ('E2', 'PRECIO APU'),
        ('F2', 'CIERRE'),
        ('G2', 'COMP.'),
        ('H2', 'MO %'),
        ('I2', 'FUENTE'),
    ]
    for celda, texto in encabezados:
        ws[celda] = texto
        ws[celda].font = bold
        ws[celda].fill = gris
        ws[celda].alignment = center
        ws[celda].border = brd

    # Datos
    fila = 3
    total_ofrecido = 0
    total_apu = 0

    for item in items_generados:
        ws.row_dimensions[fila].height = 14

        n_comp = sum(len(item.get(s, [])) for s in
                     ('materiales', 'herramientas', 'transporte', 'mano_de_obra'))

        # Calcular precio APU desde componentes
        precio_ofrecido = item.get('valor_ofrecido', 0)
        precio_apu_comp = sum(
            c['rend'] * c['unit_price']
            for s in ('materiales', 'herramientas', 'transporte', 'mano_de_obra')
            for c in item.get(s, [])
        )
        # El PRECIO UNITARIO del APU es siempre el hardcodeado = valor_ofrecido
        precio_apu = precio_ofrecido
        diferencia = abs(precio_ofrecido - precio_apu_comp)
        cierra = diferencia < 2

        # Fuente
        fuente = item.get('fuente_bd', '')
        if not fuente:
            fuente = 'Misma entidad (APU referencia)'
        if n_comp == 0:
            fuente = 'Sin componentes'

        # Porcentaje de mano de obra
        pct_mo    = item.get('pct_mo', None)
        alerta_mo = item.get('alerta_mo', False)
        mo_texto  = f'{pct_mo:.1f}%' if pct_mo is not None else '—'
        if alerta_mo:
            mo_texto = f'⚠️ {mo_texto}'

        # Color de fila: naranja si MO baja (prioridad sobre verde/ambar)
        if n_comp == 0:
            fila_fill = rojo
        elif alerta_mo:
            fila_fill = PatternFill('solid', fgColor='FCE4D6')
        elif cierra:
            fila_fill = verde
        else:
            fila_fill = ambar

        valores = [
            ('A', item.get('code', ''), center),
            ('B', item.get('description', ''), left),
            ('C', item.get('unit', ''), center),
            ('D', precio_ofrecido, right),
            ('E', precio_apu, right),
            ('F', '✅' if cierra and n_comp > 0 else ('⚠️' if n_comp > 0 else '—'), center),
            ('G', n_comp, center),
            ('H', mo_texto, center),
            ('I', fuente, left),
        ]

        for col, val, aln in valores:
            c = ws[f'{col}{fila}']
            c.value = val
            c.font = norm
            c.alignment = aln
            c.border = brd
            c.fill = fila_fill
            if col in ('D', 'E'):
                c.number_format = CUR

        total_ofrecido += precio_ofrecido
        total_apu += precio_apu
        fila += 1

    # Fila de totales
    ws.row_dimensions[fila].height = 16
    ws.merge_cells(f'A{fila}:C{fila}')
    ws[f'A{fila}'] = f'TOTAL — {fila - 3} ítems APU'
    ws[f'A{fila}'].font = bold
    ws[f'A{fila}'].fill = gris
    ws[f'A{fila}'].alignment = right
    ws[f'A{fila}'].border = brd

    for col, val in [('D', total_ofrecido), ('E', total_apu)]:
        ws[f'{col}{fila}'] = val
        ws[f'{col}{fila}'].font = bold
        ws[f'{col}{fila}'].fill = gris
        ws[f'{col}{fila}'].alignment = right
        ws[f'{col}{fila}'].border = brd
        ws[f'{col}{fila}'].number_format = CUR

    for col in ('F', 'G', 'H', 'I'):
        ws[f'{col}{fila}'].border = brd
        ws[f'{col}{fila}'].fill = gris

    # Leyenda de fuentes
    fila += 2
    naranja_fill = PatternFill('solid', fgColor='FCE4D6')
    leyenda = [
        (verde,       'Verde — APU completo con componentes (cuadra con precio ofrecido)'),
        (ambar,       'Amarillo — APU con componentes desde base de precios externa (Gobernación, INVIAS, etc.)'),
        (naranja_fill,'Naranja — APU con mano de obra inferior al 5% del costo directo — verificar tabla 2'),
        (rojo,        'Rojo — Sin componentes (ítems pendientes de completar manualmente)'),
    ]
    for fill, texto in leyenda:
        ws.row_dimensions[fila].height = 14
        ws.merge_cells(f'A{fila}:I{fila}')
        ws[f'A{fila}'] = texto
        ws[f'A{fila}'].font = Font(italic=True, name='Arial', size=8)
        ws[f'A{fila}'].fill = fill
        ws[f'A{fila}'].alignment = left
        fila += 1

def _build_apu_columnar_sheet(wb, items, factor_aiu=1.0):
    """
    Genera la hoja APU en formato columnar idéntico al de la entidad:
    CODIGO | CODINS | INSUMO | TIPO | UNIDAD | REND | UNITARIO | VR.PARCIAL | VR.TOTAL
    Fila de actividad: bold. Filas de componentes: normal.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(title='APU')

    # ── Estilos ───────────────────────────────────────────────────────────────
    thin     = Side(style='thin')
    brd      = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold9    = Font(bold=True,  name='Arial', size=9)
    norm9    = Font(bold=False, name='Arial', size=9)
    gris     = PatternFill('solid', fgColor='D9D9D9')
    azul_h   = PatternFill('solid', fgColor='BDD7EE')
    center   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_w   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    right    = Alignment(horizontal='right',  vertical='center')
    CUR      = '$ #,##0'
    REND_FMT = '#,##0.0000'

    # ── Anchos de columna (idénticos a la entidad) ────────────────────────────
    anchos = {'A':11.14,'B':7.43,'C':64.14,'D':13.86,
              'E':11.71,'F':10.0,'G':15.29,'H':15.29,'I':15.43}
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w

    # ── Fila 1: encabezados de columna ───────────────────────────────────────
    encabezados = [
        ('A','CODIGO'),('B','CODINS'),('C','INSUMO / DESCRIPCIÓN'),
        ('D','TIPO'),('E','UNIDAD'),('F','REND'),
        ('G','UNITARIO'),('H','VR. PARCIAL'),('I','VR. TOTAL'),
    ]
    ws.row_dimensions[1].height = 22
    for col, txt in encabezados:
        c = ws[f'{col}1']
        c.value     = txt
        c.font      = bold9
        c.fill      = gris
        c.alignment = center
        c.border    = brd

    # ── Datos ─────────────────────────────────────────────────────────────────
    # Mapa TIPO por sección interna
    TIPO_MAP = {
        'mano_de_obra': 'Cuadrilla',
        'herramientas': 'Equipos',
        'transporte':   'Transporte',
        'materiales':   'Insumos',
    }
    SECCIONES = ('mano_de_obra','herramientas','transporte','materiales')

    fila = 2
    for item in items:
        precio_ofrecido = item.get('valor_ofrecido', 0)

        # ── Fila de actividad (encabezado del APU) ────────────────────────────
        ws.row_dimensions[fila].height = 20
        vals_act = [
            ('A', item['code'],        center),
            ('B', '-',                 center),
            ('C', item['description'], left_w),
            ('D', 'Actividad',         center),
            ('E', item.get('unit',''), center),
            ('F', '',                  center),
            ('G', '',                  center),
            ('H', '',                  center),
            ('I', int(round(precio_ofrecido, 0)), right),
        ]
        for col, val, aln in vals_act:
            c = ws[f'{col}{fila}']
            c.value     = val
            c.font      = bold9
            c.fill      = azul_h
            c.alignment = aln
            c.border    = brd
        ws[f'I{fila}'].number_format = CUR
        fila += 1

        # ── Filas de componentes ──────────────────────────────────────────────
        for sec in SECCIONES:
            for comp in item.get(sec, []):
                if comp.get('unit_price', 0) == 0 and comp.get('rend', 0) == 0:
                    continue   # componente en cero → se omite en el output
                ws.row_dimensions[fila].height = 14
                rend  = round(comp.get('rend', 0), 6)
                up    = comp.get('unit_price', 0)
                parc  = round(rend * up, 0)
                vals_comp = [
                    ('A', item['code'],          center),
                    ('B', '',                    center),
                    ('C', comp['description'],   left_w),
                    ('D', TIPO_MAP[sec],         center),
                    ('E', comp.get('unit',''),   center),
                    ('F', rend,                  right),
                    ('G', int(round(up, 0)),     right),
                    ('H', int(parc),             right),
                    ('I', '',                    center),
                ]
                for col, val, aln in vals_comp:
                    c = ws[f'{col}{fila}']
                    c.value     = val
                    c.font      = norm9
                    c.alignment = aln
                    c.border    = brd
                ws[f'F{fila}'].number_format = REND_FMT
                ws[f'G{fila}'].number_format = CUR
                ws[f'H{fila}'].number_format = CUR
                fila += 1

    # ── Fila final: total general ─────────────────────────────────────────────
    ws.row_dimensions[fila].height = 18
    ws.merge_cells(f'A{fila}:H{fila}')
    ws[f'A{fila}'] = f'TOTAL COSTO DIRECTO — {len(items)} ítems'
    ws[f'A{fila}'].font      = bold9
    ws[f'A{fila}'].fill      = gris
    ws[f'A{fila}'].alignment = right
    ws[f'A{fila}'].border    = brd
    total_gral = sum(i.get('valor_ofrecido', 0) for i in items)
    ws[f'I{fila}'] = int(round(total_gral, 0))
    ws[f'I{fila}'].font         = bold9
    ws[f'I{fila}'].fill         = gris
    ws[f'I{fila}'].alignment    = right
    ws[f'I{fila}'].border       = brd
    ws[f'I{fila}'].number_format = CUR



def generate_apu_excel(resultado, items_manuales=None, include_aiu=False,
                       aiu_pct=0.0, bd_externas=None):
    """
    Genera el Excel final en formato columnar idéntico al de la entidad.
    Hoja 1: RESUMEN · Hoja 2: APU (columnar)
    """
    wb = Workbook()
    wb.remove(wb.active)

    # Construir lista completa de ítems aprobados
    todos = list(resultado['items_con_apu']) + (items_manuales or [])

    # Ítems sin APU que no fueron procesados manualmente
    sin_apu = resultado['items_sin_apu']
    items_sin_procesar = [i for i in sin_apu
                          if i['code'] not in {x['code'] for x in todos}]
    todos += items_sin_procesar

    if not todos:
        todos = sin_apu

    # Ordenar por código
    todos.sort(key=lambda x: _sort_key_codigo(x.get('code', '')))

    # Hoja APU columnar (formato entidad) — va primero
    _build_apu_columnar_sheet(wb, todos)

    # Hoja RESUMEN — se inserta al inicio
    _build_resumen_sheet(wb, todos)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

